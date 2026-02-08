"""
Excel Export Module for Fact-Checking Reports
Generates comprehensive Excel reports with multiple worksheets
"""

import pandas as pd
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Any, Optional
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Excel formatting will be limited.")


# Color definitions for verdicts
VERDICT_COLORS = {
    "SUPPORTED": "90EE90",           # Light green
    "REFUTED": "FFB6C1",              # Light red  
    "NOT_ENOUGH_INFO": "FFFFE0",      # Light yellow
    "INSUFFICIENT_EVIDENCE": "FFFFE0", # Light yellow (alternative name)
    "CONFLICTING_EVIDENCE": "FFD700"  # Gold
}

HEADER_COLOR = "4472C4"  # Blue
SUBHEADER_COLOR = "D9E1F2"  # Light blue


def extract_verdict_name(verdict_str: str) -> str:
    """Extract clean verdict name from various formats"""
    if 'SUPPORTED' in str(verdict_str).upper():
        return 'SUPPORTED'
    elif 'REFUTED' in str(verdict_str).upper():
        return 'REFUTED'
    elif 'CONFLICTING' in str(verdict_str).upper():
        return 'CONFLICTING_EVIDENCE'
    else:
        return 'NOT_ENOUGH_INFO'


def safe_get_attr(obj: Any, attr: str, default: Any = '') -> Any:
    """Safely get attribute from object or dict"""
    if isinstance(obj, dict):
        value = obj.get(attr, default)
    else:
        value = getattr(obj, attr, default)
    
    # Convert None to empty string for Excel
    if value is None:
        return ''
    
    return value


def generate_excel_report(
    original_input: str,
    results: Dict[str, Any],
    input_metadata: Optional[Dict[str, Any]] = None,
    metrics_summary: Optional[Dict[str, Any]] = None
) -> BytesIO:
    """
    Generate comprehensive Excel report
    
    Args:
        original_input: Original user input text
        results: Results dictionary from fact-checking
        input_metadata: Optional metadata (transcript info, language, etc.)
        metrics_summary: Optional API usage metrics summary
    
    Returns:
        BytesIO buffer containing Excel file
    """
    try:
        # Extract data from results
        is_extraction_only = results.get("mode") == "extraction_only"
        extracted_claims = results.get('extracted_claims', [])
        verified_claims = results.get('verification_results', results.get('verified_claims', []))
        preprocessed_text = results.get('preprocessed_text', '')
        
        # Create Excel writer
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Sheet 1: Summary Dashboard
            create_summary_sheet(
                writer, 
                original_input, 
                extracted_claims, 
                verified_claims,
                is_extraction_only,
                input_metadata
            )
            
            # Sheet 2: All Extracted Facts
            if extracted_claims:
                create_extracted_facts_sheet(
                    writer,
                    extracted_claims,
                    verified_claims
                )
            
            # Sheet 3: Fact-Check Results (if verified)
            if not is_extraction_only and verified_claims:
                create_fact_check_results_sheet(
                    writer,
                    verified_claims
                )
            
            # Sheet 4: Evidence Details (if verified)
            if not is_extraction_only and verified_claims:
                create_evidence_details_sheet(
                    writer,
                    verified_claims
                )
        
        # Apply formatting if openpyxl is available
        if OPENPYXL_AVAILABLE:
            buffer.seek(0)
            workbook = load_workbook(buffer)
            apply_formatting(workbook)
            
            # Save formatted workbook
            buffer = BytesIO()
            workbook.save(buffer)
        
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        raise


def create_summary_sheet(
    writer: pd.ExcelWriter,
    original_input: str,
    extracted_claims: List[Any],
    verified_claims: List[Any],
    is_extraction_only: bool,
    input_metadata: Optional[Dict[str, Any]]
):
    """Create summary dashboard sheet"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Count verdicts
    verdict_counts = {
        'SUPPORTED': 0,
        'REFUTED': 0,
        'NOT_ENOUGH_INFO': 0,
        'CONFLICTING_EVIDENCE': 0
    }
    
    if not is_extraction_only:
        for claim in verified_claims:
            verdict = extract_verdict_name(safe_get_attr(claim, 'result', ''))
            if verdict in verdict_counts:
                verdict_counts[verdict] += 1
    
    # Create summary data
    summary_data = {
        'Metric': [
            'Report Generated',
            'Input Length (characters)',
            'Total Facts Extracted',
            'Facts Verified',
            '',
            '=== VERDICT BREAKDOWN ===',
            '✅ SUPPORTED',
            '❌ REFUTED', 
            '⚠️ INSUFFICIENT INFO',
            '🔀 CONFLICTING EVIDENCE',
        ],
        'Value': [
            timestamp,
            len(original_input),
            len(extracted_claims),
            len(verified_claims) if not is_extraction_only else 'N/A',
            '',
            '',
            f"{verdict_counts['SUPPORTED']} ({verdict_counts['SUPPORTED']/max(len(verified_claims), 1)*100:.1f}%)" if verified_claims else 'N/A',
            f"{verdict_counts['REFUTED']} ({verdict_counts['REFUTED']/max(len(verified_claims), 1)*100:.1f}%)" if verified_claims else 'N/A',
            f"{verdict_counts['NOT_ENOUGH_INFO']} ({verdict_counts['NOT_ENOUGH_INFO']/max(len(verified_claims), 1)*100:.1f}%)" if verified_claims else 'N/A',
            f"{verdict_counts['CONFLICTING_EVIDENCE']} ({verdict_counts['CONFLICTING_EVIDENCE']/max(len(verified_claims), 1)*100:.1f}%)" if verified_claims else 'N/A',
        ]
    }
    
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=0)


def create_extracted_facts_sheet(
    writer: pd.ExcelWriter,
    extracted_claims: List[Any],
    verified_claims: List[Any]
):
    """Create sheet with all extracted facts"""
    data = {
        '#': [],
        'Extracted Facts': []
    }
    
    for idx, claim in enumerate(extracted_claims, 1):
        claim_text = safe_get_attr(claim, 'claim_text', str(claim))
        data['#'].append(idx)
        data['Extracted Facts'].append(claim_text)
    
    df_facts = pd.DataFrame(data)
    df_facts.to_excel(writer, sheet_name='Extracted Facts', index=False)


def create_fact_check_results_sheet(
    writer: pd.ExcelWriter,
    verified_claims: List[Any]
):
    """Create main fact-check results sheet"""
    data = {
        'Fact ID': [],
        'Fact Text': [],
        'Verdict': [],
        'Reasoning': [],
        'Corrected Claim': [],
        'Detailed Explanation': [],
        'Total Evidence Sources': [],
        'Search Queries Used': [],
        'Insufficient Info Reason': [],
        'Why We Could Not Verify': [],
        'Verification Suggestions': []
    }
    
    for idx, claim in enumerate(verified_claims, 1):
        # Extract verdict
        verdict = extract_verdict_name(safe_get_attr(claim, 'result', ''))
        
        # Extract sources - handle both list and dict formats
        sources = safe_get_attr(claim, 'sources', [])
        if not isinstance(sources, list):
            sources = []
        
        # Extract search queries - handle both list and dict formats
        search_queries = safe_get_attr(claim, 'search_queries', [])
        if not isinstance(search_queries, list):
            search_queries = []
        
        # Extract optional fields with fallbacks
        corrected_claim = safe_get_attr(claim, 'corrected_claim', '')
        
        detailed_explanation = safe_get_attr(claim, 'detailed_explanation', '')
        if not detailed_explanation or detailed_explanation == '':
            # Fallback to reasoning if no detailed explanation
            detailed_explanation = safe_get_attr(claim, 'reasoning', '')
        
        data['Fact ID'].append(idx)
        data['Fact Text'].append(safe_get_attr(claim, 'claim_text', ''))
        data['Verdict'].append(verdict)
        data['Reasoning'].append(safe_get_attr(claim, 'reasoning', ''))
        data['Corrected Claim'].append(corrected_claim)
        data['Detailed Explanation'].append(detailed_explanation)
        data['Total Evidence Sources'].append(len(sources))
        data['Search Queries Used'].append('; '.join(str(q) for q in search_queries) if search_queries else 'N/A')
        
        # Add insufficient info fields (only populated for NOT_ENOUGH_INFO verdicts)
        if verdict == 'NOT_ENOUGH_INFO':
            insufficient_reason = safe_get_attr(claim, 'insufficient_info_reason', '')
            insufficient_explanation = safe_get_attr(claim, 'insufficient_info_explanation', '')
            insufficient_suggestions = safe_get_attr(claim, 'insufficient_info_suggestions', [])
            
            # Get readable reason label
            from claim_verifier.insufficient_info_analyzer import get_reason_label
            reason_label = get_reason_label(insufficient_reason) if insufficient_reason else ''
            
            data['Insufficient Info Reason'].append(reason_label)
            data['Why We Could Not Verify'].append(insufficient_explanation)
            data['Verification Suggestions'].append(
                '\n• ' + '\n• '.join(insufficient_suggestions) if insufficient_suggestions else ''
            )
        else:
            data['Insufficient Info Reason'].append('')
            data['Why We Could Not Verify'].append('')
            data['Verification Suggestions'].append('')
    
    df_results = pd.DataFrame(data)
    df_results.to_excel(writer, sheet_name='Fact-Check Results', index=False)


def create_evidence_details_sheet(
    writer: pd.ExcelWriter,
    verified_claims: List[Any]
):
    """Create detailed evidence sheet"""
    data = {
        'Fact ID': [],
        'Fact Text': [],
        'Evidence #': [],
        'Source Title': [],
        'Source URL': []
    }
    
    for idx, claim in enumerate(verified_claims, 1):
        claim_text = safe_get_attr(claim, 'claim_text', '')
        sources = safe_get_attr(claim, 'sources', [])
        
        if not sources:
            # Add row even if no evidence
            data['Fact ID'].append(idx)
            data['Fact Text'].append(claim_text)
            data['Evidence #'].append('N/A')
            data['Source Title'].append('No evidence found')
            data['Source URL'].append('N/A')
        else:
            for ev_idx, evidence in enumerate(sources, 1):
                data['Fact ID'].append(idx)
                data['Fact Text'].append(claim_text if ev_idx == 1 else '')  # Only show on first row
                data['Evidence #'].append(ev_idx)
                data['Source Title'].append(safe_get_attr(evidence, 'title', 'Unknown'))
                data['Source URL'].append(safe_get_attr(evidence, 'url', 'N/A'))
    
    df_evidence = pd.DataFrame(data)
    df_evidence.to_excel(writer, sheet_name='Evidence Details', index=False)


def apply_formatting(workbook):
    """Apply formatting to all sheets"""
    if not OPENPYXL_AVAILABLE:
        return
    
    try:
        # Define styles
        header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Summary section colors
        summary_label_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        summary_value_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Alternating row colors
        light_row_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            
            # Format headers (first row)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Apply borders and alternating colors to all data cells
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    # Apply alternating row colors (except for Summary sheet)
                    if sheet_name != 'Summary' and row_idx % 2 == 0:
                        cell.fill = light_row_fill
            
            # Special formatting for Summary sheet
            if sheet_name == 'Summary':
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    # Column A (Metric) - light gray
                    row[0].fill = summary_label_fill
                    row[0].font = Font(bold=True, size=10)
                    # Column B (Value) - lighter gray
                    row[1].fill = summary_value_fill
                    
                    # Highlight section headers
                    if row[0].value and '===' in str(row[0].value):
                        row[0].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        row[0].font = Font(bold=True, color="FFFFFF", size=11)
                        row[1].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                # Set minimum width and cap at 100
                adjusted_width = max(min(max_length + 2, 100), 12)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Apply verdict colors in Fact-Check Results sheet
            if sheet_name == 'Fact-Check Results':
                verdict_col = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Verdict':
                        verdict_col = idx
                        break
                
                if verdict_col:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        verdict_cell = row[verdict_col - 1]
                        verdict = str(verdict_cell.value).upper()
                        
                        for verdict_key, color in VERDICT_COLORS.items():
                            if verdict_key in verdict:
                                verdict_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                verdict_cell.font = Font(bold=True)
                                break
            
            # Make URLs clickable in Evidence Details sheet
            if sheet_name == 'Evidence Details':
                url_col = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Source URL':
                        url_col = idx
                        break
                
                if url_col:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        url_cell = row[url_col - 1]
                        url = str(url_cell.value)
                        
                        if url and url.startswith('http'):
                            # Create hyperlink
                            url_cell.hyperlink = url
                            url_cell.font = Font(color="0563C1", underline="single")
                            url_cell.alignment = Alignment(vertical='top')
            
            # Freeze first row
            ws.freeze_panes = 'A2'
        
    except Exception as e:
        logger.warning(f"Could not apply formatting: {str(e)}")
